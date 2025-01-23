import openpyxl

def writedata(name, phno, regno, uid):
    try:
        wb = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb.active
        data = (name, phno, regno, uid)
        sheet.append(data)
        wb.save("Data/student_data.xlsx")
        return "Successfully Registered."
    except Exception as e:
        return f"Error: {str(e)}"

def readcardno():
    try:
        wb = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb.active
        values = [sheet.cell(row=i, column=4).value for i in range(1, sheet.max_row + 1)]
        return values
    except Exception as e:
        return f"Error: {str(e)}"

def readstudname():
    try:
        wb = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb.active
        values = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
        return values
    except Exception as e:
        return f"Error: {str(e)}"

def readregno():
    try:
        wb = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb.active
        values = [sheet.cell(row=i, column=3).value for i in range(1, sheet.max_row + 1)]
        return values
    except Exception as e:
        return f"Error: {str(e)}"

def readphno():
    try:
        wb = openpyxl.load_workbook("Data/student_data.xlsx")
        sheet = wb.active
        values = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]
        return values
    except Exception as e:
        return f"Error: {str(e)}"
